"""
question1.py

GUI application that scrapes a GitHub user's public repository names
(without using the GitHub API or python package) and saves them to an
Excel file. Includes an embedded AI chatbot (Google Gemini) that can
answer questions about the scraped repository names.

AI usage disclosure: AI assistance was used to debug a Tkinter Text
widget rendering issue (chat history not displaying after insertion)
and to identify a deprecated Gemini model name causing a 404 error.
All code was written independently.
"""
import os
import tkinter as tk

import requests     
from bs4 import BeautifulSoup
from openpyxl import Workbook
from dotenv import load_dotenv
import google.generativeai as genai


class GitHubRepoScraper:
    """
    Handles scraping a GitHub user's public repositories via HTML
    parsing (no GitHub API), saving results to Excel, and answering
    questions about the scraped repos using the Gemini API.
    """

    def __init__(self):
        load_dotenv()
        self.model_name = "gemini-2.5-flash"
        self.scraped_repos = []

    # Scraping

    def scrape_repos(self, username):
        """
        Fetch a GitHub user's repositories page and parse out repo
        names using BeautifulSoup. Returns a list of repo name
        strings, or an empty list on any network/parsing failure.
        """
        url = f"https://github.com/{username}?tab=repositories"

        try:
            response = requests.get(url)
            if response.status_code != 200:
                return []
        except requests.exceptions.RequestException:
            return []

        soup = BeautifulSoup(response.text, "html.parser")
        results = soup.find_all("a", itemprop="name codeRepository")

        repo_names = [item.text.strip() for item in results]

        self.scraped_repos = repo_names
        return repo_names

    # Excel output

    def save_to_excel(self, repos, filepath):
        """
        Write a list of repository names to an Excel file, one
        name per row under a 'Repository Name' header.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "GitHub Repositories"
        ws["A1"] = "Repository Name"

        for row_num, name in enumerate(repos, start=2):
            ws[f"A{row_num}"] = name

        wb.save(filepath)

    # Chatbot   

    def configure_api(self, api_key=None):
        """
        Configure the Gemini API with the provided key or fall back to
        the .env key.
        """
        key = api_key or os.getenv("GEMINI_API_KEY")
        if not key:
            raise ValueError("No API key provided.")
        genai.configure(api_key=key)

    def ask_about_repos(self, question):
        """
        Send a question to Gemini along with the most recently
        scraped repo names as context. Returns the model's text
        response. Raises ValueError if no repos have been scraped yet.
        """
        if not self.scraped_repos:
            raise ValueError("No repositories have been scraped yet.")

        repo_list_text = ", ".join(self.scraped_repos)
        prompt = (
            f"Here are the GitHub repositories: {repo_list_text}, "
            f"answer the following question: {question}"
        )

        model = genai.GenerativeModel(self.model_name)
        response = model.generate_content(prompt)
        return response.text


class RepoScraperGUI:
    """
    Tkinter GUI wrapping GitHubRepoScraper. Lets a user enter a
    GitHub username and a save location, scrape and export repo
    names to Excel, and ask an embedded Gemini chatbot questions
    about the scraped repos.
    """

    def __init__(self):
        self.scraper = GitHubRepoScraper()

        self.window = tk.Tk()
        self.window.title("GitHub Repository Scraper")
        self.window.geometry("550x650")

        self._build_scrape_section()
        self._build_chat_section()

    # Scrape section widgets

    def _build_scrape_section(self):
        """Create the username/folder inputs, scrape button, and status label."""
        tk.Label(self.window, text="Enter GitHub Username:").grid(
            row=0, column=0, padx=10, pady=10
        )
        self.username_entry = tk.Entry(self.window, width=30)
        self.username_entry.grid(row=0, column=1, padx=10, pady=10)

        tk.Label(self.window, text="Save Folder Location").grid(
            row=1, column=0, padx=10, pady=10
        )
        self.filepath_entry = tk.Entry(self.window, width=30)
        self.filepath_entry.grid(row=1, column=1, padx=10, pady=10)

        self.status_label = tk.Label(self.window, text="")
        self.status_label.grid(row=3, column=0, columnspan=2, pady=10)

        scrape_button = tk.Button(
            self.window, text="Scrape Repositories", command=self._on_scrape_click
        )
        scrape_button.grid(row=2, column=0, columnspan=2, pady=10)

    def _on_scrape_click(self):
        """
        Read the username and folder path from the GUI, scrape the
        user's repos, and save them to an Excel file named after
        the username in the chosen folder.
        """
        username = self.username_entry.get()
        folder_path = self.filepath_entry.get()

        repos = self.scraper.scrape_repos(username)

        if not repos:
            self.status_label.config(
                text="No repositories found or invalid username.", fg="red"
            )
            return

        full_path = os.path.join(folder_path, f"{username}_github_repos.xlsx")

        try:
            self.scraper.save_to_excel(repos, full_path)
            self.status_label.config(
                text=f"Saved {len(repos)} repos to {full_path}", fg="green"
            )
        except FileNotFoundError:
            self.status_label.config(
                text="Invalid folder path. Please check and try again.", fg="red"
            )

    # Chat section widgets

    def _build_chat_section(self):
        """Create the API key input, question entry, ask button, and chat display."""
        tk.Label(self.window, text="Ask a question:").grid(
            row=4, column=0, padx=10, pady=(10, 0), sticky="w"
        )
        self.question_entry = tk.Entry(self.window, width=40)
        self.question_entry.grid(row=5, column=0, padx=10, pady=10, sticky="w")

        ask_button = tk.Button(self.window, text="Ask", command=self._on_ask_click)
        ask_button.grid(row=5, column=1, padx=10, pady=10, sticky="w")

        tk.Label(self.window, text="Chat History:", font=("Arial", 10, "bold")).grid(
            row=6, column=0, padx=10, pady=(10, 0), sticky="w"
        )

        self.chat_history = tk.Text(self.window, height=8, width=65, wrap="word")
        self.chat_history.grid(row=7, column=0, columnspan=2, padx=10, pady=10)
        self.chat_history.config(state="disabled")
        self.chat_history.tag_config(
            "user_tag", foreground="blue", font=("Arial", 10, "bold")
        )
        self.chat_history.tag_config(
            "ai_tag", foreground="green", font=("Arial", 10, "bold")
        )

        # API key input (optional; falls back to a .env GEMINI_API_KEY if blank)
        tk.Label(self.window, text="Gemini API Key (optional):").grid(
            row=8, column=0, padx=10, pady=(10, 0), sticky="w"
        )
        self.api_key_entry = tk.Entry(self.window, width=40, show="*")
        self.api_key_entry.grid(row=8, column=1, padx=10, pady=10, sticky="w")

    def _on_ask_click(self):
        """
        Read the question (and optional API key) from the GUI, configure
        Gemini, send the question with scraped repo context, and display
        both the question and answer in the chat history.
        """
        question = self.question_entry.get()
        api_key = self.api_key_entry.get().strip()

        self.chat_history.config(state="normal")

        # configure Gemini using the typed key, or fall back to .env
        try:
            self.scraper.configure_api(api_key or None)
        except ValueError:
            self.chat_history.insert(
                tk.END, "Please provide a Gemini API key (in the box or a .env file).\n"
            )
            self.chat_history.config(state="disabled")
            return

        try:
            answer = self.scraper.ask_about_repos(question)
        except ValueError:
            self.chat_history.insert(tk.END, "Please scrape repositories first.\n")
            self.chat_history.config(state="disabled")
            return

        self.chat_history.insert(tk.END, "You: ", "user_tag")
        self.chat_history.insert(tk.END, f"{question}\n")
        self.chat_history.insert(tk.END, "Gemini: ", "ai_tag")
        self.chat_history.insert(tk.END, f"{answer}\n\n")

        self.chat_history.config(state="disabled")
        self.question_entry.delete(0, tk.END)

    # Run

    def run(self):
        """Start the Tkinter event loop."""
        self.window.mainloop()


if __name__ == "__main__":
    app = RepoScraperGUI()
    app.run()